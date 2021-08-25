import os
import csv
import re
import chardet
import xlrd
import xlwt
from xlutils.copy import copy
from shutil import copyfile
from itertools import takewhile, zip_longest
from openpyxl import load_workbook
from flask import current_app
from datetime import datetime
from werkzeug.utils import secure_filename
from lbrc_flask.database import db
from identity.model.security import User
from lbrc_flask.string_functions import similarity


# Statuses:
#
# 1 Uploaded
# 2 == column definition ==> Awaiting Submission
# 3 == submitted ==> Data Extraction
# 4 == data extracted ==> Pre-PMI Lookup
# 5 == pmi lookup complete ==> Spine Lookup
# 6 == spine lookup complete ==> Post-PMI Lookup
# 7 == pmi lookup complete ==> Create Results
# 8 == results created ==> Ready to Download
# 9 == downloaded ==> Completed
#
# Other Errors:
# 
# - Deleted
# - In Error
# - Paused
#


class DemographicsRequest(db.Model):

    id = db.Column(db.Integer, primary_key=True)
    created_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    filename = db.Column(db.String(500))
    extension = db.Column(db.String(100), nullable=False)
    owner_user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    owner = db.relationship(User, foreign_keys=[owner_user_id], backref=db.backref("demographic_requests"))
    submitted_datetime = db.Column(db.DateTime)
    deleted_datetime = db.Column(db.DateTime)
    paused_datetime = db.Column(db.DateTime)
    data_extracted_datetime = db.Column(db.DateTime)
    pmi_data_pre_completed_datetime =  db.Column(db.DateTime)
    pmi_data_post_completed_datetime =  db.Column(db.DateTime)
    lookup_completed_datetime = db.Column(db.DateTime)
    result_created_datetime = db.Column(db.DateTime)
    result_downloaded_datetime = db.Column(db.DateTime)
    error_datetime = db.Column(db.DateTime)
    error_message = db.Column(db.Text)
    last_updated_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    last_updated_by_user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    last_updated_by_user = db.relationship(User, foreign_keys=[last_updated_by_user_id])
    column_definition = db.relationship("DemographicsRequestColumnDefinition", uselist=False, back_populates="demographics_request")
    skip_pmi = db.Column(db.Boolean, default=False)

    __mapper_args__ = {
        "polymorphic_on": extension,
    }

    @property
    def filepath(self):
        return os.path.join(
            current_app.config["FILE_UPLOAD_DIRECTORY"],
            secure_filename(
                "{}_{}".format(self.last_updated_by_user.id, self.last_updated_by_user.full_name)
            ),
            secure_filename(
                "{}_{}".format(self.id, self.filename)
            ),
        )

    @property
    def result_filename(self):
        return secure_filename(
                "{}_result_{}".format(self.id, self.filename)
            )

    @property
    def result_filepath(self):
        return os.path.join(
            current_app.config["FILE_UPLOAD_DIRECTORY"],
            secure_filename(
                "{}_{}".format(self.last_updated_by_user.id, self.last_updated_by_user.full_name)
            ),
            self.result_filename,
        )

    def __lt__(self, other):
        return self.created_datetime < other.created_datetime

    @property
    def data_extracted(self):
        return self.data_extracted_datetime is not None

    @property
    def paused(self):
        return self.paused_datetime is not None

    @property
    def columns_defined(self):
        return self.column_definition and self.column_definition.is_valid

    @property
    def awaiting_submission(self):
        return self.columns_defined and self.submitted_datetime is None

    @property
    def submitted(self):
        return self.submitted_datetime is not None

    @property
    def deleted(self):
        return self.deleted_datetime is not None

    @property
    def result_created(self):
        return self.result_created_datetime is not None

    @property
    def pmi_data_pre_completed(self):
        return self.pmi_data_pre_completed_datetime is not None

    @property
    def pmi_data_post_completed(self):
        return self.pmi_data_post_completed_datetime is not None

    @property
    def result_downloaded(self):
        return self.result_downloaded_datetime is not None

    @property
    def lookup_completed(self):
        return self.lookup_completed_datetime is not None

    @property
    def in_error(self):
        return self.error_datetime is not None

    @property
    def requires_column_definition(self):
        return not self.deleted and not self.submitted and not self.in_error
    
    @property
    def requires_submission(self):
        return not self.deleted and self.awaiting_submission and not self.in_error
    
    @property
    def can_be_resubmitted(self):
        return not self.deleted and self.submitted and not self.result_created and not self.in_error
    
    @property
    def can_be_paused(self):
        return not self.deleted and self.submitted and not self.result_created and not self.paused and not self.in_error
    
    @property
    def can_be_downloaded(self):
        return not self.deleted and self.result_created and not self.in_error
    
    @property
    def can_be_deleted(self):
        return not self.deleted

    @property
    def status(self):
        if self.deleted:
            return 'Deleted'
        elif self.in_error:
            return 'Error'
        elif self.paused:
            return 'Paused'
        if not self.columns_defined:
            return 'Uploaded'
        elif not self.submitted:
            return 'Awaiting Submission'
        elif not self.data_extracted:
            return f'Extracting Data'
        elif not self.pmi_data_pre_completed and not self.skip_pmi:
            return f'Fetching PMI details {self.prepmi_count} of {self.data_count} before spine lookup'
        elif not self.lookup_completed:
            return f'Fetching Demographics {self.fetched_count} of {self.data_count}'
        elif not self.pmi_data_post_completed and not self.skip_pmi:
            return f'Fetching PMI details {self.postpmi_count} of {self.data_count} after spine lookup'
        elif not self.result_created:
            return 'Processing Demographics'
        elif not self.result_downloaded:
            return 'Ready to Download'
        else:
            return 'Downloaded'

    @property
    def data_count(self):
        return DemographicsRequestData.query.filter(DemographicsRequestData.demographics_request_id == self.id).count()

    @property
    def fetched_count(self):
        return DemographicsRequestData.query.filter(
                DemographicsRequestData.demographics_request_id == self.id
            ).filter(
                DemographicsRequestData.processed_datetime.isnot(None)
            ).count()

    @property
    def prepmi_count(self):
        return DemographicsRequestData.query.filter(
                DemographicsRequestData.demographics_request_id == self.id
            ).filter(
                DemographicsRequestData.pmi_pre_processed_datetime.isnot(None)
            ).count()

    @property
    def postpmi_count(self):
        return DemographicsRequestData.query.filter(
                DemographicsRequestData.demographics_request_id == self.id
            ).filter(
                DemographicsRequestData.pmi_post_processed_datetime.isnot(None)
            ).count()

    def get_most_likely_uhl_system_number_column_id(self):
        return self._get_most_likely_column_id('(uhl|\bs).*(number|no|)')

    def get_most_likely_nhs_number_column_id(self):
        return self._get_most_likely_column_id('nhs.*(number|no|)')

    def get_most_likely_family_name_column_id(self):
        return self._get_most_likely_column_id('(surname|(family|last).*name)')

    def get_most_likely_given_name_column_id(self):
        return self._get_most_likely_column_id('(first|given|fore).*name|name')

    def get_most_likely_gender_column_id(self):
        return self._get_most_likely_column_id('(gender|sex)')

    def get_most_likely_dob_column_id(self):
        return self._get_most_likely_column_id('(dob|date.*birth|birth.*date)')

    def get_most_likely_postcode_column_id(self):
        return self._get_most_likely_column_id('(post.*code)')

    def _get_most_likely_column_id(self, regular_expression):
        repat = re.compile(regular_expression, re.IGNORECASE)
        ids = (c.id for c in self.columns if re.search(repat, c.name))
        return next(ids, 0)

    def set_error(self, message):
        self.error_datetime = datetime.utcnow()
        self.error_message = message


class DemographicsRequestXlsx(DemographicsRequest):
    __mapper_args__ = {
        "polymorphic_identity": '.xslx',
    }

    def get_column_names(self):
        wb = load_workbook(filename=self.filepath, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=1, max_row=1)
        first_row = next(rows)

        return [c.value for c in takewhile(lambda x: x.value, first_row)]

    def iter_rows(self):
        wb = load_workbook(filename=self.filepath, read_only=True)
        ws = wb.active

        column_names = self.get_column_names()
        for r in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(column_names, r))

    def iter_result_rows(self):
        wb = load_workbook(filename=self.result_filepath)
        ws = wb.active

        rows = ws.iter_rows(min_row=1, max_row=1)
        first_row = next(rows)

        column_names = [c.value for c in takewhile(lambda x: x.value, first_row)]

        for r in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(column_names, r))

    def create_result(self):
        current_app.logger.info('DemographicsRequestXslx.create_result')
        copyfile(self.filepath, self.result_filepath)

        wb = load_workbook(filename=self.result_filepath)
        ws = wb.active

        insert_col = len(self.get_column_names()) + 1

        fieldnames = [
            'spine_nhs_number',
            'spine_title',
            'spine_forename',
            'spine_middlenames',
            'spine_lastname',
            'spine_sex',
            'spine_postcode',
            'spine_address',
            'spine_date_of_birth',
            'spine_date_of_death',
            'spine_is_deceased',
            'spine_current_gp_practice_code',
            'pmi_nhs_number',
            'pmi_uhl_system_number',
            'pmi_family_name',
            'pmi_given_name',
            'pmi_gender',
            'pmi_dob',
            'pmi_date_of_death',
            'pmi_postcode',
            'confidence',
            'spine_message',
        ]

        for i, fn in enumerate(fieldnames[::-1]):
            ws.insert_cols(insert_col)
            ws.cell(row=1, column=insert_col).value = fn

        for d in self.data:
            response = d.response

            row = d.row_number + 2

            if response:
                ws.cell(row=row, column=insert_col).value = response.nhs_number
                ws.cell(row=row, column=insert_col + 1).value = response.title
                ws.cell(row=row, column=insert_col + 2).value = response.forename
                ws.cell(row=row, column=insert_col + 3).value = response.middlenames
                ws.cell(row=row, column=insert_col + 4).value = response.lastname
                ws.cell(row=row, column=insert_col + 5).value = response.sex
                ws.cell(row=row, column=insert_col + 6).value = response.postcode
                ws.cell(row=row, column=insert_col + 7).value = response.address
                ws.cell(row=row, column=insert_col + 8).value = response.date_of_birth
                ws.cell(row=row, column=insert_col + 9).value = response.date_of_death
                ws.cell(row=row, column=insert_col + 10).value = 'True' if response.is_deceased else 'False'
                ws.cell(row=row, column=insert_col + 11).value = response.current_gp_practice_code

                pmi_data = d.pmi_data
                if pmi_data is not None:
                    ws.cell(row=row, column=insert_col + 12).value = pmi_data.nhs_number
                    ws.cell(row=row, column=insert_col + 13).value = pmi_data.uhl_system_number
                    ws.cell(row=row, column=insert_col + 14).value = pmi_data.family_name
                    ws.cell(row=row, column=insert_col + 15).value = pmi_data.given_name
                    ws.cell(row=row, column=insert_col + 16).value = pmi_data.gender
                    ws.cell(row=row, column=insert_col + 17).value = pmi_data.date_of_birth
                    ws.cell(row=row, column=insert_col + 18).value = pmi_data.date_of_death
                    ws.cell(row=row, column=insert_col + 19).value = pmi_data.postcode


            ws.cell(row=row, column=insert_col + 20).value = d.confidence
            ws.cell(row=row, column=insert_col + 21).value = '; '.join(['{} {} in {}: {}'.format(m.source, m.type, m.scope, m.message) for m in d.messages])

        wb.save(filename=self.result_filepath)


class DemographicsRequestExcel97(DemographicsRequest):
    __mapper_args__ = {
        "polymorphic_identity": '.xsl',
    }

    def get_column_names(self):
        wb = xlrd.open_workbook(filename=self.filepath)
        ws = wb.sheet_by_index(0)
        first_row = ws.row(0)

        return [c.value for c in takewhile(lambda x: x.value, first_row)]

    def iter_rows(self):
        wb = xlrd.open_workbook(filename=self.filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        column_names = self.get_column_names()
        rows = ws.get_rows()
        next(rows)
        for r in rows:
            yield dict(zip(column_names, [self._value_from_cell(c, wb.datemode) for c in r]))

    def iter_result_rows(self):
        wb = xlrd.open_workbook(filename=self.result_filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        rows = ws.get_rows()
        first_row = next(rows)

        column_names = [c.value for c in takewhile(lambda x: x.value, first_row)]

        for r in rows:
            yield dict(zip(column_names, [self._value_from_cell(c, wb.datemode) for c in r]))

    def _value_from_cell(self, cell, datemode):
        if cell.ctype == xlrd.book.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        else:
            return cell.value


    def create_result(self):
        current_app.logger.info('DemographicsRequestExcel97.create_result')

        w_book = copy(xlrd.open_workbook(filename=self.filepath, formatting_info=True))
        w_sheet = w_book.get_sheet(0)
        style = xlwt.XFStyle()
        style.num_format_str = 'D-MMM-YY' # Other options: D-MMM-YY, D-MMM, MMM-YY, h:mm, h:mm:ss, h:mm, h:mm:ss, M/D/YY h:mm, mm:ss, [h]:mm:ss, mm:ss.0

        insert_col = len(self.get_column_names())

        fieldnames = [
            'spine_nhs_number',
            'spine_title',
            'spine_forename',
            'spine_middlenames',
            'spine_lastname',
            'spine_sex',
            'spine_postcode',
            'spine_address',
            'spine_date_of_birth',
            'spine_date_of_death',
            'spine_is_deceased',
            'spine_current_gp_practice_code',
            'pmi_nhs_number',
            'pmi_uhl_system_number',
            'pmi_family_name',
            'pmi_given_name',
            'pmi_gender',
            'pmi_dob',
            'pmi_date_of_death',
            'pmi_postcode',
            'confidence',
            'spine_message',
        ]

        for i, fn in enumerate(fieldnames, start=insert_col):
            w_sheet.write(0, i, fn)

        for d in self.data:
            response = d.response

            row = d.row_number + 1

            if response:
                w_sheet.write(row, insert_col, response.nhs_number)
                w_sheet.write(row, insert_col + 1, response.title)
                w_sheet.write(row, insert_col + 2, response.forename)
                w_sheet.write(row, insert_col + 3, response.middlenames)
                w_sheet.write(row, insert_col + 4, response.lastname)
                w_sheet.write(row, insert_col + 5, response.sex)
                w_sheet.write(row, insert_col + 6, response.postcode)
                w_sheet.write(row, insert_col + 7, response.address)
                w_sheet.write(row, insert_col + 8, response.date_of_birth, style)
                w_sheet.write(row, insert_col + 9, response.date_of_death, style)
                w_sheet.write(row, insert_col + 10, 'True' if response.is_deceased else 'False')
                w_sheet.write(row, insert_col + 11, response.current_gp_practice_code)

                pmi_data = d.pmi_data
                if pmi_data is not None:
                    w_sheet.write(row, insert_col + 12, pmi_data.nhs_number)
                    w_sheet.write(row, insert_col + 13, pmi_data.uhl_system_number)
                    w_sheet.write(row, insert_col + 14, pmi_data.family_name)
                    w_sheet.write(row, insert_col + 15, pmi_data.given_name)
                    w_sheet.write(row, insert_col + 16, pmi_data.gender)
                    w_sheet.write(row, insert_col + 17, pmi_data.date_of_birth, style)
                    w_sheet.write(row, insert_col + 18, pmi_data.date_of_death, style)
                    w_sheet.write(row, insert_col + 19, pmi_data.postcode)

            w_sheet.write(row, insert_col + 20, d.confidence)
            w_sheet.write(row, insert_col + 21, '; '.join(['{} {} in {}: {}'.format(m.source, m.type, m.scope, m.message) for m in d.messages]))

        w_book.save(self.result_filepath)


class DemographicsRequestCsv(DemographicsRequest):
    __mapper_args__ = {
        "polymorphic_identity": '.csv',
    }

    def get_column_names(self):
        result = []

        with open(self.filepath, 'r', encoding=self._get_encoding()) as f:
            d_reader = csv.DictReader(f)

            result = d_reader.fieldnames

            for row in d_reader:
                for i in range(len(result), len(row)):
                    result.append(f'Column {i}')

        return result 

    def _get_encoding(self):
        rawdata = open(self.filepath, 'rb').read()
        result = chardet.detect(rawdata)
        return result['encoding']

    def iter_rows(self):
        column_names = self.get_column_names()

        with open(self.filepath, 'r', encoding=self._get_encoding()) as f:
            reader = csv.DictReader(f)

            for row in reader:
                yield {name: value for name, value in zip_longest(column_names, row.values(), fillvalue='')}

    def iter_result_rows(self):
        with open(self.result_filepath, 'r', encoding=self._get_encoding()) as f:
            reader = csv.DictReader(f)

            for row in reader:
                yield row

    def create_result(self):
        current_app.logger.info('DemographicsRequestCsv.create_result')
        with open(self.result_filepath, 'w', encoding=self._get_encoding()) as result_file:

            fieldnames = self.get_column_names()
            fieldnames.extend([
                'spine_nhs_number',
                'spine_title',
                'spine_forename',
                'spine_middlenames',
                'spine_lastname',
                'spine_sex',
                'spine_postcode',
                'spine_address',
                'spine_date_of_birth',
                'spine_date_of_death',
                'spine_is_deceased',
                'spine_current_gp_practice_code',
                'pmi_nhs_number',
                'pmi_uhl_system_number',
                'pmi_family_name',
                'pmi_given_name',
                'pmi_gender',
                'pmi_dob',
                'pmi_date_of_death',
                'pmi_postcode',
                'confidence',
                'spine_message',
            ])

            writer = csv.DictWriter(result_file, fieldnames=fieldnames)
            writer.writeheader()

            indexed_data = {d.row_number:d for d in self.data}

            for i, row in enumerate(self.iter_rows()):

                
                if i in indexed_data:
                    response = indexed_data[i].response
                    messages = indexed_data[i].messages
                    pmi_data = indexed_data[i].pmi_data
                else:
                    messages = []
                    response = None
                    pmi_data = None

                if response:
                    row['spine_nhs_number'] = response.nhs_number
                    row['spine_title'] = response.title
                    row['spine_forename'] = response.forename
                    row['spine_middlenames'] = response.middlenames
                    row['spine_lastname'] = response.lastname
                    row['spine_sex'] = response.sex
                    row['spine_postcode'] = response.postcode
                    row['spine_address'] = response.address
                    if response.date_of_birth:
                        row['spine_date_of_birth'] = response.date_of_birth.strftime('%d-%b-%Y')
                    if response.date_of_death:
                        row['spine_date_of_death'] = response.date_of_death.strftime('%d-%b-%Y')
                    row['spine_is_deceased'] = 'True' if response.is_deceased else 'False'
                    row['spine_current_gp_practice_code'] = response.current_gp_practice_code
                    row['confidence'] = indexed_data[i].confidence

                if pmi_data is not None:
                    row['pmi_nhs_number'] = pmi_data.nhs_number
                    row['pmi_uhl_system_number'] = pmi_data.uhl_system_number
                    row['pmi_family_name'] = pmi_data.family_name
                    row['pmi_given_name'] = pmi_data.given_name
                    row['pmi_gender'] = pmi_data.gender
                    row['pmi_dob'] = pmi_data.date_of_birth
                    row['pmi_date_of_death'] = pmi_data.date_of_death
                    row['pmi_postcode'] = pmi_data.postcode

                row['spine_message'] = '; '.join(['{} {} in {}: {}'.format(m.source, m.type, m.scope, m.message) for m in messages])

                writer.writerow(row)


class DemographicsRequestColumn(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    demographics_request_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequest.id))
    demographics_request = db.relationship(DemographicsRequest, foreign_keys=[demographics_request_id], backref=db.backref("columns"))
    name = db.Column(db.String(500))
    last_updated_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    last_updated_by_user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    last_updated_by_user = db.relationship(User, foreign_keys=[last_updated_by_user_id])

    def __lt__(self, other):
        return self.name < other.name
    
    def __repr__(self):
        fields = '; '.join([f'{key}="{value}"' for key, value in self.__dict__.items() if key[0] != '_' ])
        return f'[{type(self).__name__}({fields})]'


class DemographicsRequestColumnDefinition(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    demographics_request_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequest.id))
    demographics_request = db.relationship(DemographicsRequest, foreign_keys=[demographics_request_id], back_populates="column_definition")
    last_updated_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    last_updated_by_user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    last_updated_by_user = db.relationship(User, foreign_keys=[last_updated_by_user_id])

    uhl_system_number_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    uhl_system_number_column = db.relationship(DemographicsRequestColumn, foreign_keys=[uhl_system_number_column_id])
    nhs_number_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    nhs_number_column = db.relationship(DemographicsRequestColumn, foreign_keys=[nhs_number_column_id])
    family_name_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    family_name_column = db.relationship(DemographicsRequestColumn, foreign_keys=[family_name_column_id])
    given_name_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    given_name_column = db.relationship(DemographicsRequestColumn, foreign_keys=[given_name_column_id])
    gender_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    gender_column = db.relationship(DemographicsRequestColumn, foreign_keys=[gender_column_id])
    dob_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    dob_column = db.relationship(DemographicsRequestColumn, foreign_keys=[dob_column_id])
    postcode_column_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestColumn.id))
    postcode_column = db.relationship(DemographicsRequestColumn, foreign_keys=[postcode_column_id])

    gender_female_value = db.Column(db.String)
    gender_male_value = db.Column(db.String)

    @property
    def is_valid(self):
        return (
            self.nhs_number_column_id is not None
            and self.dob_column_id is not None
        ) or (
            self.family_name_column_id is not None
            and self.given_name_column_id is not None
            and self.dob_column_id is not None
            and self.gender_column_id is not None
            and self.postcode_column_id is not None
        ) or (
            self.uhl_system_number_column_id is not None
        )


class DemographicsRequestData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    demographics_request_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequest.id))
    row_number = db.Column(db.Integer, nullable=False)
    demographics_request = db.relationship(DemographicsRequest, backref=db.backref("data"))
    response = db.relationship("DemographicsRequestDataResponse", uselist=False, back_populates="demographics_request_data")
    pmi_data = db.relationship("DemographicsRequestPmiData", uselist=False, back_populates="demographics_request_data")

    nhs_number = db.Column(db.String)
    uhl_system_number = db.Column(db.String)
    family_name = db.Column(db.String)
    given_name = db.Column(db.String)
    gender = db.Column(db.String)
    dob = db.Column(db.String)
    postcode = db.Column(db.String)

    created_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    processed_datetime = db.Column(db.DateTime)
    pmi_pre_processed_datetime = db.Column(db.DateTime)
    pmi_post_processed_datetime = db.Column(db.DateTime)

    @property
    def processed(self):
        return self.processed_datetime is not None

    @property
    def pmi_pre_processed(self):
        return self.pmi_pre_processed_datetime is not None

    @property
    def pmi_post_processed(self):
        return self.pmi_post_processed_datetime is not None

    @property
    def has_error(self):
        return any(m.is_error for m in self.messages)

    @property
    def has_error(self):
        return any(m.is_error for m in self.messages)

    @property
    def confidence(self):
        if self.response is None:
            return 0
        
        parts = []

        if self.nhs_number:
            parts.append(similarity(self.nhs_number, self.response.nhs_number))
        if self.family_name:
            parts.append(similarity(self.family_name, self.response.lastname))
        if self.given_name:
            parts.append(similarity(self.given_name, self.response.forename))
        if self.postcode:
            parts.append(similarity(self.postcode, self.response.postcode))

        return round(sum(parts) / len(parts), 2)

    
    def __repr__(self):
        fields = '; '.join([f'{key}="{value}"' for key, value in self.__dict__.items() if key[0] != '_' ])
        return f'[{type(self).__name__}({fields})]'


class DemographicsRequestPmiData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    demographics_request_data_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestData.id))
    demographics_request_data = db.relationship(DemographicsRequestData, foreign_keys=[demographics_request_data_id], back_populates="pmi_data")

    nhs_number = db.Column(db.String)
    uhl_system_number = db.Column(db.String)
    family_name = db.Column(db.String)
    given_name = db.Column(db.String)
    gender = db.Column(db.String)
    date_of_birth = db.Column(db.Date)
    date_of_death = db.Column(db.Date)
    postcode = db.Column(db.String)

    created_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    def __eq__(self, other):
        if isinstance(other, DemographicsRequestPmiData):
            return (
                self.nhs_number == other.nhs_number and
                self.uhl_system_number == other.uhl_system_number and
                self.family_name == other.family_name and
                self.given_name == other.given_name and
                self.gender == other.gender and
                self.date_of_birth == other.date_of_birth and
                self.date_of_death == other.date_of_death and
                self.postcode == other.postcode
            )
        return False

    def __ne__(self, other):
        return not self.__eq__(other)


class DemographicsRequestDataMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    demographics_request_data_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestData.id))
    demographics_request_data = db.relationship(DemographicsRequestData, backref=db.backref("messages"))

    type = db.Column(db.String)
    source = db.Column(db.String)
    scope = db.Column(db.String)
    message = db.Column(db.String)
    created_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    @property
    def is_error(self):
        return self.type == 'error'


class DemographicsRequestDataResponse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    demographics_request_data_id = db.Column(db.Integer, db.ForeignKey(DemographicsRequestData.id))
    demographics_request_data = db.relationship(DemographicsRequestData, foreign_keys=[demographics_request_data_id], back_populates="response")
    created_datetime = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    nhs_number = db.Column(db.String)
    title = db.Column(db.String)
    forename = db.Column(db.String)
    middlenames = db.Column(db.String)
    lastname = db.Column(db.String)
    sex = db.Column(db.String)
    postcode = db.Column(db.String)
    address = db.Column(db.String)
    date_of_birth = db.Column(db.Date)
    date_of_death = db.Column(db.Date)
    is_deceased = db.Column(db.Boolean)
    current_gp_practice_code = db.Column(db.String)
